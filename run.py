from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

import requests
from bs4 import BeautifulSoup as soup

############### xlsx.opener ###############


def open_xls():
    wb = load_workbook(filename='ADD_py.xlsx')
    ws = wb.get_sheet_by_name('A1')
    return ws


def get_row_content(ws, i):
    my_list = []

    for cell in ws[i]:
        my_list.append(cell.value)
    # print(my_list)
    # print(my_list[0])
    return my_list
###################### decoder ######################


def decoder(cod):
    return(cod.decode('cp1251')).encode('utf8')

########### selenium ####################
# вот здесь она лагает - не всегда догружается элемент(кнопка), в итоге кликать нечего и всё летит к чертям:)
def get_page(drug_name, drug_dosage):
    browser = webdriver.Chrome(executable_path='C:\Program Files (x86)\Google\Chrome\Application\chromedriver')
    browser.get("https://tabletki.ua/list/")
    wait = WebDriverWait(browser, 30)
    search_bar = browser.find_element_by_id("ctl00_ctl00_HeaderPanel_QueryPanel_QueryCtrl")
    search_bar.send_keys(drug_name)
    search_button = browser.find_element_by_id('ctl00_ctl00_HeaderPanel_QueryPanel_SubmitLink')
    search_button.click()
    inst_button = browser.find_element_by_xpath('//*[@id="smart-menu"]/ul/li[1]/a')
    inst_button.click()
    banner = wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="top-banner-panel"]')))
    dosage_button = wait.until(ec.element_to_be_clickable(
        (By.XPATH, '//*[@id="ctl00_HeaderPanel_QueryPanel_ControlsPanel"]/div[1]/div[1]')))
    dosage_button.click()
    dropdown_list = browser.find_element_by_xpath('//*[@id="ctl00_HeaderPanel_QueryPanel_ControlsPanel"]/div[1]/div[1]/ul')
    if dropdown_list.is_displayed():
        pass
    else:
        wait.until(ec.element_to_be_clickable(dosage_button))
        dosage_button.click()
    dropdown_list = browser.find_element_by_xpath\
        ('//*[@id="ctl00_HeaderPanel_QueryPanel_ControlsPanel"]/div[1]/div[1]/ul')
    option = wait.until(ec.element_to_be_clickable((By.XPATH, '//*[contains(text(), "%s")]' % drug_dosage)))
    option.click()
    # print(my_page)
    return my_page

######################## Page parsing ######################


def get_page_content(page):
    # page = browser.current_url
    # URL = 'https://tabletki.ua/%D0%A6%D0%B8%D0%B1%D0%BE%D1%80/25903/'
    r = requests.get('%s' % page)
    page = r.content
    page_soup = soup(page, "html.parser")
    # print(page_soup.prettify())
    drug_data = page_soup.find_all('div', {"id": "ctl00_MAIN_ContentPlaceHolder_NeedTranslatePanel"})
    return drug_data
    # print(drug_data)

def get_item_content(drug_data):
    my_text = []
    for item in drug_data:
        # print(item.contents[1].find_all(['h2']))
        for header in item.contents[1].find_all(['h2']):
            my_text.append(header.get_text())
            for elem in header.next_elements:
                if elem.name and elem.name.startswith('h'):
                    break
                if elem.name == 'p':
                    my_text.append(elem.get_text())
    a = '\n'.join(my_text)
    return a


####################### main ######################

def main():
    ws = open_xls()
    # get_current_row(ws, i=1)
    i = 2
    while i in range(2, ws.max_row + 1):
        my_list = get_row_content(ws, i)
        # print(my_list)
        drug_id, drug_name, drug_dosage = my_list[0], my_list[1], my_list[2]
        my_page = get_page(drug_name, drug_dosage)
        drug_data = get_page_content(page=my_page)
        my_text = get_item_content(drug_data)
        # print(my_text)

        def load_to_doc(d):
            doc = Document()
            title = '%s інструкція для застосування' % drug_name
            doc.add_heading('{}'.format(title), level=2)
            doc.add_paragraph(str(d))
            doc.save('{}.docx'.format(drug_id))
        load_to_doc(d=my_text)

        i += 1
    print("Completed!")


if __name__ == "__main__":
     main()
